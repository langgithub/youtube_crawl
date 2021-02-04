#!/usr/bin/env python3
# -*- coding:utf-8 -*- 
# author：yuanlang 
# creat_time: 2021/1/20 下午2:04
# file: excel_format.py


from openpyxl import load_workbook, Workbook


def update_to_excel(path):
    wb = load_workbook(path)
    sheets = wb.worksheets  # 获取当前所有的sheet

    sheet1 = sheets[0]
    print(sheet1)

    rows = sheet1.max_row
    for i in range(1, rows + 1):
        status = sheet1.cell(row=i, column=4).value
        if status =="yes":
            url = sheet1.cell(row=i, column=5).value
            if url is not None and ".mp4" in url:
                new_url = url.replace(".mp4", "")
                sheet1.cell(row=i, column=5, value=new_url)

    wb.save(filename=path)


if __name__ == "__main__":
    pass
    # update_to_excel("youtube_创意广告.xlsx")
    # update_to_excel("youtube_泰国广告.xlsx")