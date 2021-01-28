#!/usr/bin/env python3
# -*- coding:utf-8 -*- 
# author：yuanlang 
# creat_time: 2021/1/18 上午10:39
# file: ad_crawl.py

# -*- coding: utf-8 -*-
import os
import io
import json
import base64
import pickle
import requests
import urllib3
import functools
import random
import time
import hashlib
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from scrapy import Selector
from abc import ABCMeta, abstractmethod
from PIL import Image, ImageChops
from concurrent.futures import ThreadPoolExecutor
from multiprocessing import Process
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from you_get.extractors import (
    imgur,
    magisto,
    youtube,
    missevan,
    acfun,
    bilibili,
    soundcloud,
    tiktok
)

from obssdk.obs import ObsOperator, MultipartUploadFileRequest
from obssdk.obs.util import *
from io import BytesIO, BufferedReader
from redis_cluster_helper import redis_cluster



urllib3.disable_warnings()
# 深圳地址
obs_host = "obs-cn-shenzhen.yun.pingan.com"
obs_access_key = 'QTU2QjkyRjFFRDhGNDJDMzlBRkFDQUUyMTU5Q0Y5REM'
obs_secret_key = 'QzFBODBEQTFDNEVENEU1NDlDQkEwNjMwQzk2MzU2MkM'
bucket_name = "spider"


class ErrorLog(object):

    def __init__(self, bg, full, diff, offset):
        self.bg = bg
        self.full = full
        self.diff = diff
        self.offset = offset


class BaseSpiderCrack(metaclass=ABCMeta):
    """验证码破解基础类"""

    def __init__(self, driver):
        self.driver = driver
        self.driver.maximize_window()
        self.error: ErrorLog = None

    @abstractmethod
    def crack(self):
        pass


class AdCrawl(BaseSpiderCrack):
    """youtube 广告爬虫"""

    def __init__(self, driver):
        super().__init__(driver)

    def check_response(self):
        """检查是否成功"""
        pickle.dump(self.driver.get_cookies(), open("cookies.pkl", "wb"))

    def crack(self):
        """执行破解程序"""
        # self.write_to_excel("youtube.xlsx", [[1,2],[3,2]])

        self.driver.get("https://www.youtube.com/results?search_query=%E5%88%9B%E6%84%8F%E5%B9%BF%E5%91%8A")

        while 1:
            time.sleep(5)
            js = 'document.documentElement.scrollTop=1000000000000;return document.documentElement.scrollTop;'
            scrollTop = self.driver.execute_script(js)
            print(scrollTop)
            if "无更多结果" in self.driver.page_source:
                break

        print("over")
        html = Selector(text=self.driver.page_source)
        videos = html.css("div#container a#video-title")
        result_videos = []
        for video in videos:
            title = video.css("a::attr(title)").extract()[0]
            href = "https://www.youtube.com" + video.css("a::attr(href)").extract()[0]
            label = video.css("a::attr(aria-label)").extract()[0]
            result_videos.append([title, href, label])

        self.write_to_excel("youtube.xlsx", result_videos)

    def write_to_excel(self, path, items):
        wb = load_workbook(path)
        # wb = Workbook()
        sheets = wb.worksheets  # 获取当前所有的sheet

        sheet1 = sheets[0]
        print(sheet1)

        # 通过Cell对象读取
        # rows = sheet1.max_row
        # column = sheet1.max_column
        # print(rows, column)

        i = 0
        for rows in items:
            i = i + 1
            j = 0
            for value in rows:
                j = j + 1
                sheet1.cell(row=i, column=j, value=value)
        wb.save(path)


def update_to_excel(path, items_key):
    wb = load_workbook(path)
    # wb = Workbook()
    sheets = wb.worksheets  # 获取当前所有的sheet

    sheet1 = sheets[0]
    print(sheet1)

    rows = sheet1.max_row
    for i in range(1, rows + 1):
        title = sheet1.cell(row=i, column=1).value
        if title in items_key:
            sheet1.cell(row=i, column=4, value="yes")
            sheet1.cell(row=i, column=5, value=items_key[title])
    wb.save(filename=path)


def video_list():
    """视频列表下载"""
    from selenium.webdriver.chrome.options import Options
    options = Options()
    # 我本地的 Google Chrome
    options.binary_location = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'
    # 我本地的 chromedriver
    driver = webdriver.Chrome(chrome_options=options, executable_path='/Users/yuanlang/work/javascript/chromedriver')
    # 这个js没用到。过滑块可以使用
    with open('stealth.min.js') as f:
        js = f.read()

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": js
    })

    cracker = AdCrawl(driver)
    cracker.crack()


def get_href(path):
    wb = load_workbook(path)
    sheets = wb.worksheets  # 获取当前所有的sheet
    sheet1 = sheets[0]
    # 通过Cell对象读取
    rows = sheet1.max_row
    column = sheet1.max_column
    print(rows, column)
    return [sheet1.cell(row=i, column=2).value for i in range(1, rows + 1) if
            sheet1.cell(row=i, column=4).value == "no"]


def video_download(limit, _file_name):
    # rh = redis_cluster()
    # _ip = "{}:31289".format(rh.rpoplpush("hk_vps", "hk_vps"))
    ip_list = []
    try:
        response = requests.get("https://lookout.pais.pingan.com/proxy_server/select?isown=3&protocol=2&site=test&token=4cc5fbe69e2a93d48bef68319b763541&count={0}".format(limit))
        ip_list = json.loads(response.text)["data"]
    except Exception as e:
        print(e)

    def download(_command):
        try:
            result = os.system(_command)
            # youtube.download(url, merge=True, output_dir='video_dir', itag=18)
            print("download url>>>> {0} 下载完成 result={1}".format(url, result))
        except Exception as e:
            print(e)

    href1 = get_href(_file_name)
    # href1 = get_href("youtube_创意广告.xlsx")
    # href2 = get_href("youtube_泰国广告.xlsx")
    # hrefs = href1 + href2
    hrefs = href1[:limit]

    # with ThreadPoolExecutor(max_workers=5) as t:
    #     for href in hrefs:
    #         t.submit(download, href)
    pools = []
    for href in hrefs:
        _ip = ip_list.pop()
        command = "you-get -o /root/project/youtube_crawl/video_dir -x {} --itag=18 {}".format(_ip, href)
        print("excute command >>>>> {0}".format(command))
        p = Process(target=download, args=(command,))
        pools.append(p)

    for p in pools:
        p.start()

    for p in pools:
        p.join()

    print("本轮下载结束")
    return len(hrefs) == 0


def notify_upload(upload_id, state, total_parts, finish_parts):
    print("%s %s %s of %s" % (upload_id, state, finish_parts, total_parts))


def video_upload(_file_name):
    def upload(_path, _name):
        obs = ObsOperator(obs_host, obs_access_key, obs_secret_key)
        _file_path = "{}/video_dir/{}".format(_path, _name)
        size = int(os.path.getsize(_file_path) / float(1024 * 1024))
        _input_name = hashlib.md5()  # 要加密的字符串
        _input_name.update(_name.encode("utf-8"))
        _url = "http://obs-cn-shenzhen.yun.pingan.com/{0}/{1}.mp4".format(bucket_name, _input_name.hexdigest())
        print(_url)
        if size < 100:
            ret = obs.put_object_from_file(bucket_name, _input_name.hexdigest(), _file_path)
            print(ret.get_e_tag())
        else:
            multipart_request = MultipartUploadFileRequest()
            multipart_request.set_bucket_name(bucket_name)
            multipart_request.set_object_key(_input_name.hexdigest())
            multipart_request.set_upload_file_path(_file_path)
            multipart_request.set_upload_notifier(notify_upload)
            obs.put_object_multipart(multipart_request)

    # path = os.path.dirname(__file__)
    path = "/root/project/youtube_crawl"
    dir_path = path + "/video_dir"
    print(dir_path)
    dirs = os.listdir(dir_path)
    look_urls = {}
    pools = []
    for name in dirs:
        p = Process(target=upload, args=(path, name))
        pools.append(p)
        input_name = hashlib.md5()  # 要加密的字符串
        input_name.update(name.encode("utf-8"))
        url = "http://obs-cn-shenzhen.yun.pingan.com/{0}/{1}.mp4".format(bucket_name, input_name.hexdigest())
        look_urls[name.replace(".mp4", "")] = url

    for p in pools:
        p.start()

    for p in pools:
        p.join()

    update_to_excel(_file_name, look_urls)
    # update_to_excel("youtube_创意广告.xlsx", look_urls)
    # update_to_excel("youtube_泰国广告.xlsx", look_urls)
    for name in dirs:
        file_path = "{}/video_dir/{}".format(path, name)
        os.remove(file_path)


def get_ip():
    _ip = "127.0.0.1"
    try:
        import socket
        host = socket.gethostname()
        _ip = socket.gethostbyname(host)
    except Exception as e:
        print(e)
    return _ip


if __name__ == "__main__":
    ip = get_ip()
    flag = False
    is_running = True
    file_name = ""
    if ip == "127.0.0.1":
        # is_running = False
        file_name = "youtube_创意广告.xlsx"
    elif ip == "10.20.29.130":
        file_name = "youtube_创意广告.xlsx"
    elif ip == "10.20.29.131":
        file_name = "youtube_泰国广告.xlsx"
    elif ip == "10.8.184.25":
        file_name = "youtube_创意广告.xlsx"

    # run
    while is_running:
        flag = video_download(5, file_name)
        video_upload(file_name)
        if flag: break
